import streamlit as st
import pandas as pd
import re
from io import BytesIO
import zipfile
import os
from openpyxl import load_workbook
import tempfile
from typing import Dict, List, Tuple, Optional
from collections import defaultdict

def parse_course_semester_grade_from_column(column_name: str) -> Optional[Tuple[str, str, str, str]]:
    """
    Parse a column name in format 'Course-Semester-Year-Grade' or similar variations.
    
    Args:
        column_name: String like 'MATH101-Fall2024-A' or 'BIO202-Spring2025-B+'
    
    Returns:
        Tuple of (course, semester, year, grade) or None if parsing fails
    """
    # Pattern to match course-semester-year-grade format
    # Supports various grade formats including A+, B-, etc.
    pattern = r'^([A-Z]+\d+)-([A-Za-z]+)(\d{4})-([A-F][+-]?|[A-F]|[A-Z][+-]?)$'
    
    match = re.match(pattern, column_name.strip())
    if match:
        course, semester, year, grade = match.groups()
        return course, semester, year, grade
    
    # Alternative pattern for different formats
    pattern2 = r'^([A-Z]+\d+)[-_]([A-Za-z]+)[-_](\d{4})[-_]([A-F][+-]?|[A-F]|[A-Z][+-]?)$'
    match2 = re.match(pattern2, column_name.strip())
    if match2:
        course, semester, year, grade = match2.groups()
        return course, semester, year, grade
    
    return None

def parse_course_semester_grade_from_value(value: str) -> Optional[Tuple[str, str, str, str]]:
    """
    Parse course information from cell value in format 'Course/Semester-Year/Grade'.
    
    Args:
        value: String like 'SPTH201/FALL-2016/F' or 'MATH101/SPRING-2024/A+'
    
    Returns:
        Tuple of (course, semester, year, grade) or None if parsing fails
    """
    if pd.isna(value) or not isinstance(value, str):
        return None
        
    value = value.strip()
    if not value:
        return None
    
    # Pattern for format: COURSE/SEMESTER-YEAR/GRADE
    pattern1 = r'^([A-Z]+\d+[A-Z]*)/([A-Z]+-\d{4})/([A-F][+-]?|[A-Z][+-]?|P\*?|R)$'
    match1 = re.match(pattern1, value.upper())
    if match1:
        course, semester_year, grade = match1.groups()
        # Split semester-year
        if '-' in semester_year:
            semester, year = semester_year.split('-', 1)
            return course, semester, year, grade
    
    # Alternative pattern: COURSE/SEMESTER/YEAR/GRADE
    pattern2 = r'^([A-Z]+\d+[A-Z]*)/([A-Z]+)/(\d{4})/([A-F][+-]?|[A-Z][+-]?|P\*?|R)$'
    match2 = re.match(pattern2, value.upper())
    if match2:
        course, semester, year, grade = match2.groups()
        return course, semester, year, grade
    
    # Pattern for incomplete entries (missing grade)
    pattern3 = r'^([A-Z]+\d+[A-Z]*)/([A-Z]+-\d{4})/?$'
    match3 = re.match(pattern3, value.upper())
    if match3:
        course, semester_year = match3.groups()
        if '-' in semester_year:
            semester, year = semester_year.split('-', 1)
            return course, semester, year, "INCOMPLETE"
    
    return None

def identify_grade_columns(df: pd.DataFrame) -> list:
    """
    Identify columns that contain grade data based on naming patterns or content.
    
    Args:
        df: Input DataFrame
    
    Returns:
        List of column names that appear to contain grade data
    """
    grade_columns = []
    
    # First, check if column names follow the expected pattern
    for col in df.columns:
        if parse_course_semester_grade_from_column(str(col)) is not None:
            grade_columns.append(col)
    
    # If no columns match the naming pattern, look for columns that contain course data
    if not grade_columns:
        for col in df.columns:
            if col.upper().startswith('COURSE'):
                # Check if this column contains valid course data
                sample_values = df[col].dropna().head(5)
                for val in sample_values:
                    if parse_course_semester_grade_from_value(str(val)) is not None:
                        grade_columns.append(col)
                        break
    
    return grade_columns

def transform_grades_to_tidy(df: pd.DataFrame) -> pd.DataFrame:
    """
    Transform wide-format grade data to tidy format.
    
    Args:
        df: Input DataFrame with wide-format grade data
    
    Returns:
        Transformed DataFrame in tidy format
    """
    # Make a copy to avoid modifying original
    df_copy = df.copy()
    
    # Drop columns that are entirely empty
    df_copy = df_copy.dropna(axis=1, how='all')
    
    # Identify demographic/ID columns (non-grade columns)
    grade_columns = identify_grade_columns(df_copy)
    id_columns = [col for col in df_copy.columns if col not in grade_columns]
    
    if not grade_columns:
        st.warning("No grade columns detected. Please ensure your column names follow the format: Course-Semester-Year-Grade (e.g., MATH101-Fall2024-A)")
        return pd.DataFrame()
    
    # Melt the DataFrame to convert from wide to long format
    melted_df = pd.melt(
        df_copy,
        id_vars=id_columns,
        value_vars=grade_columns,
        var_name='Course_Semester_Grade',
        value_name='Grade'
    )
    
    # Filter out missing values
    melted_df = melted_df.dropna(subset=['Grade'])
    
    # Remove rows where Grade is empty string or whitespace
    melted_df = melted_df[melted_df['Grade'].astype(str).str.strip() != '']
    
    # Parse the Course_Semester_Grade column (which contains actual course data, not column names)
    parsed_data = []
    
    for _, row in melted_df.iterrows():
        # First try parsing as column name format
        parsed = parse_course_semester_grade_from_column(str(row['Course_Semester_Grade']))
        
        # If that fails, try parsing as cell value format
        if not parsed:
            parsed = parse_course_semester_grade_from_value(str(row['Grade']))
        
        if parsed:
            course, semester, year, grade = parsed
            
            # Create new row with parsed data
            new_row = {}
            
            # Add ID columns
            for id_col in id_columns:
                new_row[id_col] = row[id_col]
            
            # Add parsed course information
            new_row['Course'] = course
            new_row['Semester'] = semester.title()  # Convert to title case (Fall, Spring, etc.)
            new_row['Year'] = int(year)
            new_row['Grade'] = grade
            
            parsed_data.append(new_row)
    
    # Create final DataFrame
    if parsed_data:
        tidy_df = pd.DataFrame(parsed_data)
        
        # Reorder columns for better readability
        base_columns = [col for col in id_columns]
        course_columns = ['Course', 'Semester', 'Year', 'Grade']
        final_columns = base_columns + course_columns
        
        tidy_df = tidy_df[final_columns]
        
        return tidy_df
    else:
        return pd.DataFrame()

# ===== INTERNSHIP DATA PROCESSING FUNCTIONS =====

def extract_student_id(file_path: str) -> Optional[str]:
    """
    Extract student ID from 'Current Semester Advising' sheet, cell C5
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Student ID as string or None if not found/error
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        
        # Check if 'Current Semester Advising' sheet exists
        if 'Current Semester Advising' not in workbook.sheetnames:
            return None
            
        sheet = workbook['Current Semester Advising']
        student_id = sheet['C5'].value
        
        if student_id is not None:
            return str(student_id).strip()
        return None
        
    except Exception as e:
        st.error(f"Error extracting student ID: {str(e)}")
        return None

def extract_internship_data(file_path: str) -> Optional[Dict[str, int]]:
    """
    Extract internship data (completed hours by internship code) from Excel file
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dictionary mapping internship codes to completed hours, or None if error
    """
    try:
        # Try to read all sheets and find the one with internship data
        excel_file = pd.ExcelFile(file_path)
        
        internship_data = {}
        
        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                
                # Look for the header pattern: Internship Code, Total Hours, Completed, Remaining
                for i in range(len(df)):
                    row = df.iloc[i]
                    if (len(row) >= 4 and 
                        pd.notna(row.iloc[0]) and 
                        pd.notna(row.iloc[2]) and
                        str(row.iloc[0]).strip().lower() == 'internship code' and
                        str(row.iloc[2]).strip().lower() == 'completed'):
                        
                        # Found the header row, now extract data
                        for j in range(i + 1, len(df)):
                            data_row = df.iloc[j]
                            if (len(data_row) >= 4 and 
                                pd.notna(data_row.iloc[0]) and 
                                pd.notna(data_row.iloc[2])):
                                
                                internship_code = str(data_row.iloc[0]).strip()
                                completed_hours = data_row.iloc[2]
                                
                                # Skip if internship code is empty or completed hours is not numeric
                                if internship_code and pd.notna(completed_hours):
                                    try:
                                        completed_hours = int(float(completed_hours))
                                        internship_data[internship_code] = completed_hours
                                    except (ValueError, TypeError):
                                        continue
                            else:
                                # Stop when we hit empty rows
                                break
                        
                        # Found data in this sheet, return it
                        if internship_data:
                            return internship_data
                            
            except Exception:
                continue  # Try next sheet
                
        return internship_data if internship_data else None
        
    except Exception as e:
        st.error(f"Error extracting internship data: {str(e)}")
        return None

def process_zip_file(uploaded_file) -> Tuple[pd.DataFrame, List[str], List[str]]:
    """
    Process uploaded zip file and extract student internship data
    
    Args:
        uploaded_file: Streamlit uploaded file object
        
    Returns:
        Tuple of (consolidated_dataframe, processed_files, error_files)
    """
    processed_files = []
    error_files = []
    all_student_data = []
    
    with tempfile.TemporaryDirectory() as temp_dir:
        # Extract zip file
        with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
            
            # Get all Excel files from the extracted directory
            excel_files = []
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file.endswith(('.xlsx', '.xls')):
                        excel_files.append(os.path.join(root, file))
            
            if not excel_files:
                st.error("No Excel files found in the uploaded zip file.")
                return pd.DataFrame(), [], []
            
            # Process each Excel file
            for file_path in excel_files:
                file_name = os.path.basename(file_path)
                
                try:
                    # Extract student ID
                    student_id = extract_student_id(file_path)
                    if not student_id:
                        error_files.append(f"{file_name}: Could not extract student ID from 'Current Semester Advising' sheet, cell C5")
                        continue
                    
                    # Extract internship data
                    internship_data = extract_internship_data(file_path)
                    if not internship_data:
                        error_files.append(f"{file_name}: Could not extract internship data")
                        continue
                    
                    # Add student data to consolidated list
                    student_record = {'Student_ID': student_id, **internship_data}
                    all_student_data.append(student_record)
                    
                    processed_files.append(file_name)
                    
                except Exception as e:
                    error_files.append(f"{file_name}: {str(e)}")
    
    # Create consolidated DataFrame
    if all_student_data:
        consolidated_df = pd.DataFrame(all_student_data)
        # Fill NaN values with 0 for internship codes not present in all files
        consolidated_df = consolidated_df.fillna(0)
        # Ensure Student_ID is the first column
        cols = ['Student_ID'] + [col for col in consolidated_df.columns if col != 'Student_ID']
        consolidated_df = consolidated_df[cols]
    else:
        consolidated_df = pd.DataFrame()
    
    return consolidated_df, processed_files, error_files

# ===== ADVISING STATUS EXTRACTOR FUNCTIONS =====

def process_pbhl_advising(uploaded_zip) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Process PBHL advising sheets and return summary and conflict-free groups."""

    summary = defaultdict(lambda: {"Yes": 0, "Optional": 0, "Not Advised": 0})
    student_yes = {}

    with zipfile.ZipFile(uploaded_zip) as z:
        for file_name in z.namelist():
            if not file_name.lower().endswith(('.xlsx', '.xls')):
                continue
            with z.open(file_name) as f:
                df = pd.read_excel(f, sheet_name="Current Semester Advising", header=None)
                df = df.iloc[7:, [0, 7]]  # Columns: Course Code, Advising Status
                df.columns = ["Course Code", "Advised"]
                df = df.dropna(subset=["Course Code"])  # Drop rows without course code
                df["Advised"] = df["Advised"].fillna("").astype(str).str.strip().str.title()

                # Update summary counts
                for _, row in df.iterrows():
                    course = str(row["Course Code"]).strip()
                    status = row["Advised"] if row["Advised"] in ["Yes", "Optional"] else "Not Advised"
                    summary[course][status] += 1

                # Collect courses advised as "Yes" for conflict-free groups
                yes_courses = sorted(df[df["Advised"] == "Yes"]["Course Code"].astype(str).str.strip().tolist())
                student_name = os.path.splitext(os.path.basename(file_name))[0]
                student_yes[student_name] = yes_courses

    # Create summary DataFrame
    summary_df = pd.DataFrame(
        [
            {
                "Course Code": course,
                "Yes Count": counts["Yes"],
                "Optional Count": counts["Optional"],
                "Not Advised Count": counts["Not Advised"],
            }
            for course, counts in sorted(summary.items())
        ]
    )

    # Create conflict-free groups DataFrame
    groups = defaultdict(list)
    for student, courses in student_yes.items():
        groups[tuple(courses)].append(student)

    max_courses = max((len(k) for k in groups.keys()), default=0)
    group_rows = []
    for courses, students in groups.items():
        row = {"Students": ", ".join(sorted(students))}
        for i, course in enumerate(courses, start=1):
            row[f"Course {i}"] = course
        group_rows.append(row)

    group_df = pd.DataFrame(group_rows)
    for i in range(1, max_courses + 1):
        col = f"Course {i}"
        if col not in group_df.columns:
            group_df[col] = ""
    group_df = group_df[["Students"] + [f"Course {i}" for i in range(1, max_courses + 1)]]

    return summary_df, group_df


def process_spth_advising(uploaded_zip) -> Tuple[pd.DataFrame, pd.DataFrame, bytes]:
    """Process SPTH advising sheets: clean files, summarize, and group courses."""

    courses = [
        "BIOL 201", "BSE-000", "SPTH 205", "SPTH 206", "SPTH 229", "ARAB 201", "BCOM 300",
        "BIOL 201 or CHEM 201", "CIVL 201", "CIVL 202", "CMPS 202", "COMM 201", "ENGL 201",
        "ENGL 202", "PSYCH 201", "STAT 201", "SPTH 201", "SPTH 202", "SPTH 204", "SPTH 205N",
        "SPTH 206N", "SPTH 207", "SPTH 208", "SPTH 209", "SPTH 210", "SPTH 211", "SPTH 212",
        "SPTH 217", "SPTH 219", "SPTH 224", "SPTH 225", "SPTH 226", "SPTH 227", "SPTH 228",
        "SPTH 229N", "SPTH 230", "SPTH 231", "SPTH 234", "SPTH 240", "SPTH 240S", "SPTH 241",
        "SPTH 241S", "SPTH 250", "SPTH 270", "SPTH 271", "SPTH 272", "SPTH 273", "SPTH 274",
        "SPTH 275", "SPTH 276", "SPTH 277", "SPTH 290", "SPTH 291", "SPTH 292", "SPTH 293A",
        "SPTH 293B", "SPTH 294A", "SPTH 294B", "SPTH 296A", "SPTH 296B", "SPTH 297A", "SPTH 297B",
        "SPTH 298A", "SPTH 298B", "SPTH 299A", "SPTH 299B", "INEG 200", "INEG 300", "MATH 101",
        "MATH 102", "CHEM 101", "BIOL 101", "PHYS 101", "ENGL 101"
    ]

    summary = {course: {"Yes": 0, "Optional": 0, "Not Advised": 0} for course in courses}
    student_yes = {}

    processed_buffer = BytesIO()

    with zipfile.ZipFile(uploaded_zip) as zin, zipfile.ZipFile(processed_buffer, 'w') as zout:
        for file_name in zin.namelist():
            if not file_name.lower().endswith(('.xlsx', '.xls')):
                continue
            with zin.open(file_name) as f:


                student_name = os.path.splitext(os.path.basename(file_name))[0]
                yes_courses = []

                for course in courses:
                    row = df[df["Course Code"] == course]
                    if row.empty:
                        status = "Not Advised"
                    else:
                        val = str(row["Advised"].iloc[0]).strip().title()
                        status = val if val in ["Yes", "Optional"] else "Not Advised"
                    summary[course][status] += 1
                    if status == "Yes":
                        yes_courses.append(course)

                student_yes[student_name] = yes_courses

    processed_buffer.seek(0)

    # Summary DataFrame
    summary_df = pd.DataFrame(
        [
            {
                "Course Code": course,
                "Yes Count": counts["Yes"],
                "Optional Count": counts["Optional"],
                "Not Advised Count": counts["Not Advised"],
            }
            for course, counts in summary.items()
        ]
    )

    # Conflict-free groups
    groups = defaultdict(list)
    for student, courses_taken in student_yes.items():
        groups[tuple(courses_taken)].append(student)

    max_courses = max((len(k) for k in groups.keys()), default=0)
    group_rows = []
    for courses_taken, students in groups.items():
        row = {"Students": ", ".join(sorted(students))}
        for i, course in enumerate(courses_taken, start=1):
            row[f"Course {i}"] = course
        group_rows.append(row)

    group_df = pd.DataFrame(group_rows)
    for i in range(1, max_courses + 1):
        col = f"Course {i}"
        if col not in group_df.columns:
            group_df[col] = ""
    group_df = group_df[["Students"] + [f"Course {i}" for i in range(1, max_courses + 1)]]

    return summary_df, group_df, processed_buffer.getvalue()


def advising_status_extractor_tab():
    """Streamlit interface for the Advising Status Extractor tool."""
    st.header("🗂 Advising Status Extractor")
    track = st.selectbox("Select Program", ["PBHL", "SPTH"])
    uploaded_zip = st.file_uploader("Upload ZIP of advising sheets", type=["zip"])

    if uploaded_zip is None:
        return

    if track == "PBHL":
        summary_df, group_df = process_pbhl_advising(uploaded_zip)
        st.subheader("Advising Summary")
        st.dataframe(summary_df, use_container_width=True)
        st.subheader("Conflict-Free Course Groups")
        st.dataframe(group_df, use_container_width=True)
    else:
        summary_df, group_df, processed_zip = process_spth_advising(uploaded_zip)
        st.download_button(
            "Download Processed Files",
            data=processed_zip,
            file_name="processed_advising_files.zip",
        )
        st.subheader("Advising Summary")
        st.dataframe(summary_df, use_container_width=True)
        st.subheader("Conflict-Free Course Groups")
        st.dataframe(group_df, use_container_width=True)

# ===== TAB FUNCTIONS =====

def grade_transformer_tab():
    st.header("📊 Grade Data Transformer")
    st.markdown("Transform wide-format student grade Excel files into tidy, normalized data for academic analysis")
    
    # File upload section
    st.subheader("1. Upload Excel File")
    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=['xlsx', 'xls'],
        help="Upload an Excel file containing student grade data in wide format"
    )
    
    if uploaded_file is not None:
        try:
            # Read the Excel file
            df = pd.read_excel(uploaded_file)
            
            st.success(f"✅ File uploaded successfully! Shape: {df.shape[0]} rows × {df.shape[1]} columns")
            
            # Show original data preview
            st.subheader("2. Original Data Preview")
            st.dataframe(df.head(10), use_container_width=True)
            
            # Analyze the data
            st.subheader("3. Data Analysis")
            
            # Check for empty columns
            empty_cols = df.columns[df.isnull().all()].tolist()
            if empty_cols:
                st.info(f"Found {len(empty_cols)} empty columns that will be removed: {', '.join(empty_cols)}")
            
            # Identify grade columns
            grade_columns = identify_grade_columns(df)
            if grade_columns:
                st.info(f"Detected {len(grade_columns)} grade columns: {', '.join(grade_columns[:5])}{'...' if len(grade_columns) > 5 else ''}")
            else:
                st.error("❌ No grade columns detected. Please ensure your column names follow the format: Course-Semester-Year-Grade (e.g., MATH101-Fall2024-A)")
                return
            
            # Transform the data
            st.subheader("4. Data Transformation")
            
            with st.spinner("Transforming data..."):
                tidy_df = transform_grades_to_tidy(df)
            
            if not tidy_df.empty:
                st.success(f"✅ Data transformed successfully! New shape: {tidy_df.shape[0]} rows × {tidy_df.shape[1]} columns")
                
                # Show transformation summary
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Original Rows", df.shape[0])
                with col2:
                    st.metric("Transformed Rows", tidy_df.shape[0])
                with col3:
                    st.metric("Unique Students", tidy_df.iloc[:, 0].nunique() if len(tidy_df.columns) > 0 else 0)
                
                # Show transformed data preview
                st.subheader("5. Transformed Data Preview")
                st.dataframe(tidy_df.head(20), use_container_width=True)
                
                # Show summary statistics
                st.subheader("6. Summary Statistics")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.subheader("Courses")
                    if 'Course' in tidy_df.columns:
                        course_counts = tidy_df['Course'].value_counts()
                        st.dataframe(course_counts.head(10), use_container_width=True)
                
                with col2:
                    st.subheader("Grade Distribution")
                    if 'Grade' in tidy_df.columns:
                        grade_counts = tidy_df['Grade'].value_counts()
                        st.dataframe(grade_counts, use_container_width=True)
                
                # Download section
                st.subheader("7. Download Cleaned Data")
                
                # Create Excel file in memory
                output = BytesIO()
                tidy_df.to_excel(output, engine='openpyxl', sheet_name='Cleaned_Data', index=False)
                
                output.seek(0)
                
                st.download_button(
                    label="📥 Download Cleaned Excel File",
                    data=output.getvalue(),
                    file_name="cleaned_student_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            else:
                st.error("❌ No valid data could be extracted. Please check your file format and column naming conventions.")
                
        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
            st.info("Please ensure your file is a valid Excel file with the correct format.")
    
    # Instructions section
    with st.expander("📋 Instructions & Format Requirements"):
        st.markdown("""
        ### Expected File Format
        
        Your Excel file should have:
        
        1. **Student Information Columns**: ID, NAME, CURRICULUM, etc.
        2. **Grade Columns**: Can be in two formats:
           
           **Format 1** - Column names: `Course-Semester-Year-Grade`
           - Example: `MATH101-Fall2024-A`, `BIO202-Spring2025-B+`
           
           **Format 2** - Course columns (COURSE_1, COURSE_2, etc.) with cell values: `Course/Semester-Year/Grade`
           - Example: `SPTH201/FALL-2016/F`, `MATH101/SPRING-2024/A+`
           - Course: Letters and numbers (e.g., SPTH201, MATH101)
           - Semester: FALL, SPRING, SUMMER, WINTER
           - Year: 4-digit year (e.g., 2016, 2024)
           - Grade: Letter grades (A, B, C, D, F, P, R) with optional +/- modifiers
        
        ### What This Tool Does
        
        1. **Loads** your Excel file into a DataFrame
        2. **Drops** any columns that are entirely empty
        3. **Melts** the data from wide to long format
        4. **Filters** out missing values
        5. **Splits** combined Course_Semester_Grade strings into separate fields:
           - Course (e.g., "MATH101")
           - Semester (e.g., "Fall")
           - Year (e.g., "2024")
           - Grade (e.g., "A")
        6. **Exports** the result as a new Excel file
        
        ### Output Format
        
        The cleaned data will have columns like:
        ```
        ID | NAME | MAJOR | Course | Semester | Year | Grade
        ```
        """)

def internship_consolidator_tab():
    st.header("🎓 Internship Data Consolidator")
    st.markdown("Upload a zip file containing Excel files with student internship data. Extract student IDs and consolidate completed internship hours into a single report.")
    
    # File upload section
    st.subheader("Upload Zip File")
    uploaded_file = st.file_uploader(
        "Choose a zip file containing student Excel files",
        type=['zip'],
        help="Upload a zip file containing Excel files (.xlsx) with student internship data"
    )
    
    if uploaded_file is not None:
        st.success(f"Uploaded: {uploaded_file.name}")
        
        # Process button
        if st.button("Process Files", type="primary"):
            with st.spinner("Processing files... This may take a few moments."):
                # Process the zip file
                consolidated_df, processed_files, error_files = process_zip_file(uploaded_file)
                
                # Display results
                st.subheader("Processing Results")
                
                # Summary statistics
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Files Processed Successfully", len(processed_files))
                with col2:
                    st.metric("Files with Errors", len(error_files))
                with col3:
                    st.metric("Total Students", len(consolidated_df) if not consolidated_df.empty else 0)
                
                # Display errors if any
                if error_files:
                    st.subheader("⚠️ Files with Errors")
                    with st.expander("View Error Details"):
                        for error in error_files:
                            st.error(error)
                
                # Display processed files
                if processed_files:
                    st.subheader("✅ Successfully Processed Files")
                    with st.expander("View Processed Files"):
                        for file in processed_files:
                            st.success(file)
                
                # Display consolidated data
                if not consolidated_df.empty:
                    st.subheader("Consolidated Data Preview")
                    st.dataframe(
                        consolidated_df,
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    # Download section
                    st.subheader("Download Consolidated Report")
                    
                    # Convert DataFrame to Excel bytes
                    output = BytesIO()
                    consolidated_df.to_excel(output, engine='openpyxl', sheet_name='Consolidated_Report', index=False)
                    
                    excel_data = output.getvalue()
                    
                    st.download_button(
                        label="📥 Download Consolidated Excel Report",
                        data=excel_data,
                        file_name="consolidated_internship_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    # Display summary information
                    st.subheader("Report Summary")
                    st.info(f"""
                    - **Total Students**: {len(consolidated_df)}
                    - **Internship Codes Found**: {len([col for col in consolidated_df.columns if col != 'Student_ID'])}
                    - **Files Successfully Processed**: {len(processed_files)}
                    - **Files with Errors**: {len(error_files)}
                    """)
                else:
                    st.error("No data could be extracted from the uploaded files. Please check the file format and ensure they contain the required sheets and data structure.")
    
    # Instructions section
    with st.expander("📋 Instructions and Requirements"):
        st.markdown("""
        ### File Requirements:
        - Upload a **zip file** containing Excel files (.xlsx format)
        - Each Excel file should represent one student's data
        - Each file must have a sheet named **"Current Semester Advising"** with the student ID in cell **C5**
        - Each file should contain internship data with columns: **Internship Code**, **Total Hours**, **Completed**, **Remaining**
        
        ### Expected Data Format:
        The application looks for internship data in the following format:
        ```
        Internship Code | Total Hours | Completed | Remaining
        SPTH290        |     50      |     25    |    25
        SPTH291        |     50      |     30    |    20
        ...
        ```
        
        ### Output:
        - Consolidated Excel file with student IDs as rows and internship codes as columns
        - Values represent completed hours for each internship per student
        - Missing internship codes for a student will show as 0
        """)

def main():
    st.set_page_config(
        page_title="PU PBHL Department Tools",
        page_icon="🎓",
        layout="wide"
    )
    
    st.title("🎓 PU PBHL Department Academic Data Tools")
    st.markdown("Comprehensive toolkit for processing student academic data")
    
    # Create tabs
    tab1, tab2, tab3 = st.tabs([
        "📊 Grade Data Transformer",
        "🎓 Internship Data Consolidator",
        "🗂 Advising Status Extractor",
    ])

    with tab1:
        grade_transformer_tab()

    with tab2:
        internship_consolidator_tab()

    with tab3:
        advising_status_extractor_tab()

if __name__ == "__main__":
    main()
