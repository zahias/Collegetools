import streamlit as st
import pandas as pd
from io import BytesIO
import zipfile, os, tempfile
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional

def extract_student_meta(path: str) -> Tuple[Optional[str], Optional[str]]:
    """Return the student ID and name found on the Current Semester Advising sheet."""

    def _strip(value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        value = str(value).strip()
        return value or None

    def _neighbor_value(sheet, row: int, col: int) -> Optional[str]:
        # Prefer same-row values first so names appear next to their label.
        for r, c in ((row, col + 1), (row + 1, col)):
            if r <= sheet.max_row and c <= sheet.max_column:
                val = _strip(sheet.cell(row=r, column=c).value)
                if val:
                    return val
        return None

    try:
        wb = load_workbook(path, data_only=True)
        if "Current Semester Advising" not in wb.sheetnames:
            return None, None
        sh = wb["Current Semester Advising"]

        sid = _strip(sh["C5"].value)
        name = _strip(sh["C4"].value)

        # Fall back to label search close to the header area where the metadata typically lives.
        if not name:
            local_labels = {"student name", "name"}
            for row in sh.iter_rows(min_row=1, max_row=15, max_col=10):
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.strip().lower() in local_labels:
                        name = _neighbor_value(sh, cell.row, cell.column)
                        if name:
                            break
                if name:
                    break

        # Broader sweep that only reacts to labels containing both "student" and "name".
        if not name:
            for row in sh.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    normalized = val.strip().lower()
                    if "student" in normalized and "name" in normalized:
                        name = _neighbor_value(sh, cell.row, cell.column)
                        if name:
                            break
                if name:
                    break

        return sid, name
    except Exception:
        return None, None

def extract_internship_data(path: str) -> Optional[Dict[str, int]]:
    try:
        xls = pd.ExcelFile(path)
        out = {}
        for sh in xls.sheet_names:
            try:
                df = pd.read_excel(path, sheet_name=sh, header=None)
                for i in range(len(df)):
                    row = df.iloc[i]
                    if (len(row) >= 4 and
                        pd.notna(row.iloc[0]) and pd.notna(row.iloc[2]) and
                        str(row.iloc[0]).strip().lower() == "internship code" and
                        str(row.iloc[2]).strip().lower() == "completed"):
                        for j in range(i+1, len(df)):
                            r = df.iloc[j]
                            if len(r) >= 4 and pd.notna(r.iloc[0]) and pd.notna(r.iloc[2]):
                                code = str(r.iloc[0]).strip()
                                try:
                                    comp = int(float(r.iloc[2]))
                                except Exception:
                                    continue
                                if code: out[code] = comp
                            else:
                                break
                        if out: return out
            except Exception:
                continue
        return out if out else None
    except Exception:
        return None

def process_zip(up_zip) -> Tuple[pd.DataFrame, List[str], List[str]]:
    processed, errors, all_rows = [], [], []
    with tempfile.TemporaryDirectory() as td:
        with zipfile.ZipFile(up_zip, "r") as zf:
            zf.extractall(td)
        excel_files = []
        for root,_,files in os.walk(td):
            for f in files:
                if f.endswith((".xlsx",".xls")):
                    excel_files.append(os.path.join(root,f))
        if not excel_files:
            return pd.DataFrame(), [], ["No Excel files found in the zip."]
        for path in excel_files:
            file_name = os.path.basename(path)
            sid, student_name = extract_student_meta(path)
            data = extract_internship_data(path)
            if not sid:
                errors.append(f"{file_name}: missing Student ID at 'Current Semester Advising'!C5")
                continue
            if not data:
                errors.append(f"{file_name}: internship table not found")
                continue
            row = {"Student ID": sid, "Student Name": student_name or "", **data}
            all_rows.append(row)
            processed.append(file_name)

    if not all_rows:
        return pd.DataFrame(), processed, errors

    df = pd.DataFrame(all_rows).fillna(0)
    preferred = ["Student ID", "Student Name"]
    cols = preferred + [c for c in df.columns if c not in preferred]
    return df[cols], processed, errors

def run():
    st.subheader("🎓 Internship Data Consolidator")
    up = st.file_uploader("Upload a .zip of Excel files (one student per file)", type=["zip"])

    if up and st.button("Process", type="primary"):
        with st.spinner("Reading files…"):
            df, ok, bad = process_zip(up)
        c1,c2,c3 = st.columns(3)
        with c1: st.metric("Processed", len(ok))
        with c2: st.metric("Errors", len(bad))
        with c3: st.metric("Students", 0 if df.empty else len(df))
        if bad:
            st.markdown("**Errors**")
            for e in bad: st.error(e)
        if not df.empty:
            st.markdown("**Consolidated Preview**")
            st.dataframe(df, use_container_width=True, hide_index=True)
            out = BytesIO(); df.to_excel(out, engine="openpyxl", index=False, sheet_name="Consolidated_Report")
            st.download_button("📥 Download Excel", out.getvalue(),
                               file_name="consolidated_internship_report.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
