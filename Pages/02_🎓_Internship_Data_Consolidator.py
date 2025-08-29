import streamlit as st
import pandas as pd
from io import BytesIO
import zipfile, os, tempfile
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional

st.set_page_config(page_title="Internship Data Consolidator", page_icon="🎓", layout="wide")

def extract_student_id(file_path: str) -> Optional[str]:
    try:
        wb = load_workbook(file_path, data_only=True)
        if "Current Semester Advising" not in wb.sheetnames:
            return None
        sheet = wb["Current Semester Advising"]
        sid = sheet["C5"].value
        return str(sid).strip() if sid is not None else None
    except Exception:
        return None

def extract_internship_data(file_path: str) -> Optional[Dict[str, int]]:
    try:
        xls = pd.ExcelFile(file_path)
        out = {}
        for sh in xls.sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sh, header=None)
                for i in range(len(df)):
                    row = df.iloc[i]
                    if (len(row) >= 4 and
                        pd.notna(row.iloc[0]) and pd.notna(row.iloc[2]) and
                        str(row.iloc[0]).strip().lower() == "internship code" and
                        str(row.iloc[2]).strip().lower() == "completed"):
                        for j in range(i+1, len(df)):
                            r = df.iloc[j]
                            if len(r) >= 4 and pd.notna(r.iloc[0]) and pd.notna(r.iloc[2]):
                                code = str(r.iloc[0]).strip()
                                try:
                                    comp = int(float(r.iloc[2]))
                                except Exception:
                                    continue
                                if code:
                                    out[code] = comp
                            else:
                                break
                        if out:
                            return out
            except Exception:
                continue
        return out if out else None
    except Exception:
        return None

def process_zip(uploaded_zip) -> Tuple[pd.DataFrame, List[str], List[str]]:
    processed, errors, all_rows = [], [], []
    with tempfile.TemporaryDirectory() as td:
        with zipfile.ZipFile(uploaded_zip, "r") as zf:
            zf.extractall(td)
        excel_files = []
        for root, _, files in os.walk(td):
            for f in files:
                if f.endswith((".xlsx", ".xls")):
                    excel_files.append(os.path.join(root, f))
        if not excel_files:
            return pd.DataFrame(), [], ["No Excel files found in the zip."]

        for path in excel_files:
            name = os.path.basename(path)
            sid = extract_student_id(path)
            data = extract_internship_data(path)
            if not sid:
                errors.append(f"{name}: missing Student ID at 'Current Semester Advising'!C5")
                continue
            if not data:
                errors.append(f"{name}: internship table not found")
                continue
            row = {"Student_ID": sid, **data}
            all_rows.append(row)
            processed.append(name)

    if not all_rows:
        return pd.DataFrame(), processed, errors

    df = pd.DataFrame(all_rows).fillna(0)
    cols = ["Student_ID"] + [c for c in df.columns if c != "Student_ID"]
    return df[cols], processed, errors

# --------------- UI ----------------
st.title("🎓 Internship Data Consolidator")
up = st.file_uploader("Upload a .zip of Excel files (one student per file)", type=["zip"])

if up and st.button("Process", type="primary"):
    with st.spinner("Reading files…"):
        df, ok, bad = process_zip(up)

    c1, c2, c3 = st.columns(3)
    with c1: st.metric("Processed", len(ok))
    with c2: st.metric("Errors", len(bad))
    with c3: st.metric("Students", 0 if df.empty else len(df))

    if bad:
        st.subheader("Errors")
        for e in bad:
            st.error(e)

    if not df.empty:
        st.subheader("Consolidated Preview")
        st.dataframe(df, use_container_width=True, hide_index=True)

        out = BytesIO()
        df.to_excel(out, engine="openpyxl", index=False, sheet_name="Consolidated_Report")
        st.download_button("📥 Download Excel", out.getvalue(),
                           file_name="consolidated_internship_report.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
